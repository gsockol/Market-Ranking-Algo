"""
src/exporter.py
===============
Writes a multi-sheet Excel workbook (.xlsx) using openpyxl.

Sheets
------
1. Rankings          — composite score, tier, rank, category contributions
2. Raw Data          — all raw values per country (CSV + API + manual)
3. Derived Metrics   — penetration headroom, dues, opportunity, etc.
4. Normalised Scores — 0–1 normalised value per variable per country
5. Weight Matrix     — per-country weight for each variable (reflects Rules 1-3)
6. Data Sources      — audit trail (source label per variable per country)

Formatting
----------
- Header row: dark background, white bold text
- Tier rows: colour-coded by tier
- Score column: conditional colour gradient
- Percentage columns: % number format
- Currency columns: $#,##0.00 format
- Frozen header rows and auto-fitted column widths
"""

import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Font, PatternFill, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL   = PatternFill("solid", fgColor="0F172A")
_HEADER_FONT   = Font(color="F1F5F9", bold=True, size=10)
_ALT_FILL      = PatternFill("solid", fgColor="F8FAFC")
_TIER_FILLS = {
    1: PatternFill("solid", fgColor="DCFCE7"),   # green-100
    2: PatternFill("solid", fgColor="DBEAFE"),   # blue-100
    3: PatternFill("solid", fgColor="FEF3C7"),   # amber-100
    4: PatternFill("solid", fgColor="FEE2E2"),   # red-100
}
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="E2E8F0")
)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)

_VAR_LABELS = {
    "opportunity_usd_m":        "Opportunity ($M)",
    "potential_market_size":    "Potential Market Size ($M)",
    "gym_membership_cagr":      "Gym Membership CAGR",
    "penetration_headroom":     "Penetration Headroom",
    "concentration":            "Concentration (000s/gym)",
    "ease_of_doing_business":   "Ease of Doing Business",
    "political_stability":      "Political Stability",
    "inflation_rate":           "Inflation Rate",
    "currency_volatility":      "Currency Volatility",
    "rule_of_law":              "Rule of Law",
    "financing_accessibility": "Ease of Financing (GFDD)",
    "corporate_tax_rate":       "Corporate Tax Rate",
    "labor_cost_index":         "Labour Cost Index",
    "real_estate_cost_index":   "Real Estate Cost Index",
    "youth_population_pct":     "Youth Population % (15–34)",
    "middle_class_pct":         "Middle Class %",
    "avg_gym_spend_pct_gdp":    "Avg Gym Spend as % of GDP",
}

_DERIVED_COLS = [
    "penetration_headroom",
    "implied_members_current",
    "current_dues_monthly_usd",
    "dues_increase_pct",
    "future_dues_monthly_usd",
    "implied_members_future",
    "potential_market_size",
    "opportunity_usd_m",
    "avg_gym_spend_pct_gdp",
]


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _tier_number(tier_str: str) -> int:
    for t in [1, 2, 3, 4]:
        if str(t) in tier_str:
            return t
    return 4


def _write_header(ws, headers: list, row: int = 1):
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _autofit(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            ws.column_dimensions[col_letter].width = max(
                min_width, min(max_width, max(lengths) + 2)
            )


def _safe(val):
    """Convert NaN / None to empty string for Excel."""
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    return val


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _sheet_rankings(wb, scores_df, categories):
    ws = wb.create_sheet("Rankings")
    cat_contrib_cols = [f"contrib_{k}" for k in categories]
    cat_labels = {f"contrib_{k}": v for k, v in {
        "contrib_market_opportunity":   "Market Opp. (pts)",
        "contrib_penetration_headroom": "Penetration (pts)",
        "contrib_operational_risk":     "Op. Risk (pts)",
        "contrib_cost_structure":       "Cost (pts)",
        "contrib_demand_indicators":    "Demand (pts)",
    }.items()}

    headers = ["Rank", "Country", "Composite Score", "Tier"] + [
        cat_labels.get(c, c) for c in cat_contrib_cols
    ]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    for r_idx, row in scores_df.iterrows():
        excel_row = r_idx + 2
        tier_n = _tier_number(str(row["tier"]))
        fill = _TIER_FILLS.get(tier_n, _ALT_FILL)

        ws.cell(excel_row, 1, value=row["rank"])
        ws.cell(excel_row, 2, value=row["country"])
        score_cell = ws.cell(excel_row, 3, value=round(row["composite_score"], 2))
        score_cell.number_format = "0.00"
        ws.cell(excel_row, 4, value=row["tier"])

        for c_idx, cat_col in enumerate(cat_contrib_cols, 5):
            val = row.get(cat_col, 0.0)
            cell = ws.cell(excel_row, c_idx, value=round(_safe(val) or 0, 2))
            cell.number_format = "0.00"

        for col in range(1, len(headers) + 1):
            ws.cell(excel_row, col).fill = fill
            ws.cell(excel_row, col).border = _THIN_BORDER
            ws.cell(excel_row, col).font = _NORMAL_FONT

    _autofit(ws)


def _sheet_raw_data(wb, full_df, scored_vars):
    ws = wb.create_sheet("Raw Data")
    display_vars = [v for v in scored_vars if v in full_df.columns]
    headers = ["Country"] + [_VAR_LABELS.get(v, v) for v in display_vars]
    _write_header(ws, headers)
    ws.freeze_panes = "B2"

    for r_idx, row in full_df.iterrows():
        excel_row = r_idx + 2
        ws.cell(excel_row, 1, value=row["country"]).font = _BOLD_FONT
        for c_idx, var in enumerate(display_vars, 2):
            val = _safe(row.get(var))
            cell = ws.cell(excel_row, c_idx, value=val if val != "" else None)
            cell.font = _NORMAL_FONT
            cell.border = _THIN_BORDER
            if isinstance(val, float):
                cell.number_format = "#,##0.00"

    _autofit(ws)


def _sheet_derived(wb, full_df):
    ws = wb.create_sheet("Derived Metrics")
    avail_cols = [c for c in _DERIVED_COLS if c in full_df.columns]
    pretty = {
        "penetration_headroom":      "Penetration Headroom",
        "implied_members_current":   "Implied Members (Current)",
        "current_dues_monthly_usd":  "Current Monthly Dues (USD)",
        "dues_increase_pct":         "Dues Increase Rate",
        "future_dues_monthly_usd":   "Future Monthly Dues (USD)",
        "implied_members_future":    "Implied Members (Future)",
        "potential_market_size":     "Potential Market Size ($M)",
        "opportunity_usd_m":         "Opportunity ($M)",
        "avg_gym_spend_pct_gdp":     "Avg Gym Spend as % of GDP",
    }
    headers = ["Country"] + [pretty.get(c, c) for c in avail_cols]
    _write_header(ws, headers)
    ws.freeze_panes = "B2"

    for r_idx, row in full_df.iterrows():
        excel_row = r_idx + 2
        ws.cell(excel_row, 1, value=row["country"]).font = _BOLD_FONT
        for c_idx, col in enumerate(avail_cols, 2):
            val = _safe(row.get(col))
            cell = ws.cell(excel_row, c_idx, value=val if val != "" else None)
            cell.font = _NORMAL_FONT
            cell.border = _THIN_BORDER
            if isinstance(val, float):
                cell.number_format = "#,##0.0000"

    _autofit(ws)


def _sheet_normalised(wb, normalized_df, scored_vars, full_df):
    ws = wb.create_sheet("Normalised Scores")
    headers = ["Country"] + [_VAR_LABELS.get(v, v) for v in scored_vars]
    _write_header(ws, headers)
    ws.freeze_panes = "B2"

    countries = full_df["country"].tolist()
    for r_idx, country in enumerate(countries):
        excel_row = r_idx + 2
        ws.cell(excel_row, 1, value=country).font = _BOLD_FONT
        norm_row = normalized_df.iloc[r_idx] if r_idx < len(normalized_df) else pd.Series()
        for c_idx, var in enumerate(scored_vars, 2):
            val = _safe(norm_row.get(var, np.nan))
            cell = ws.cell(excel_row, c_idx, value=val if val != "" else None)
            cell.font = _NORMAL_FONT
            cell.border = _THIN_BORDER
            if isinstance(val, float):
                cell.number_format = "0.000"

    _autofit(ws)


def _sheet_weights(wb, weight_matrix, scored_vars, base_weights):
    ws = wb.create_sheet("Weight Matrix")
    headers = ["Country"] + [_VAR_LABELS.get(v, v) for v in scored_vars]
    _write_header(ws, headers)

    # Base weights row
    base_row = ws.max_row + 1
    ws.cell(base_row, 1, value="[BASE WEIGHTS]").font = Font(bold=True, italic=True, size=10)
    for c_idx, var in enumerate(scored_vars, 2):
        cell = ws.cell(base_row, c_idx, value=round(base_weights.get(var, 0.0), 4))
        cell.number_format = "0.0%"
        cell.font = Font(italic=True, size=10, color="64748B")

    ws.freeze_panes = "B2"

    for r_idx, (country, weights) in enumerate(weight_matrix.items()):
        excel_row = r_idx + 3
        ws.cell(excel_row, 1, value=country).font = _BOLD_FONT
        for c_idx, var in enumerate(scored_vars, 2):
            w = weights.get(var, 0.0)
            base_w = base_weights.get(var, 0.0)
            cell = ws.cell(excel_row, c_idx, value=round(w, 4))
            cell.number_format = "0.0%"
            cell.border = _THIN_BORDER
            if abs(w - base_w) > 1e-6:
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
                cell.font = Font(bold=True, size=10, color="92400E")
            elif w == 0.0:
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
                cell.font = Font(size=10, color="991B1B")
            else:
                cell.font = _NORMAL_FONT

    _autofit(ws)


def _sheet_sources(wb, audit, scored_vars, full_df):
    ws = wb.create_sheet("Data Sources")
    headers = ["Country"] + [_VAR_LABELS.get(v, v) for v in scored_vars]
    _write_header(ws, headers)
    ws.freeze_panes = "B2"

    source_colors = {
        "csv_derived":   "DBEAFE",
        "api":           "D1FAE5",
        "manual_yaml":   "FEF3C7",
        "manual_prompt": "FEF3C7",
        "missing":       "FEE2E2",
    }
    source_labels = {
        "csv_derived":   "CSV / Derived",
        "api":           "World Bank API",
        "manual_yaml":   "Manual (YAML)",
        "manual_prompt": "Manual (Prompt)",
        "missing":       "MISSING",
    }

    countries = full_df["country"].tolist()
    for r_idx, country in enumerate(countries):
        excel_row = r_idx + 2
        ws.cell(excel_row, 1, value=country).font = _BOLD_FONT
        country_audit = audit.get(country, {})
        for c_idx, var in enumerate(scored_vars, 2):
            src = country_audit.get(var, "missing")
            label = source_labels.get(src, src)
            color = source_colors.get(src, "FFFFFF")
            cell = ws.cell(excel_row, c_idx, value=label)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(size=9)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    _autofit(ws, min_width=12, max_width=18)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def export_excel(
    scores_df: pd.DataFrame,
    full_df: pd.DataFrame,
    normalized_df: pd.DataFrame,
    weight_matrix: dict,
    audit: dict,
    categories: dict,
    base_weights: dict,
    output_dir: str,
    filename: str,
) -> str:
    """
    Write the multi-sheet Excel workbook and return its path.
    """
    scored_vars = list(base_weights.keys())

    out_path = Path(output_dir)
    out_path.mkdir(exist_ok=True)
    filepath = out_path / filename

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _sheet_rankings(wb, scores_df, categories)
    _sheet_raw_data(wb, full_df, scored_vars)
    _sheet_derived(wb, full_df)
    _sheet_normalised(wb, normalized_df, scored_vars, full_df)
    _sheet_weights(wb, weight_matrix, scored_vars, base_weights)
    _sheet_sources(wb, audit, scored_vars, full_df)

    wb.save(str(filepath))
    return str(filepath)
